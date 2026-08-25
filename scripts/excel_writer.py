"""Writes extracted Schedule H data to an Excel workbook, one worksheet per
source page, named after that page's page number.
"""
from __future__ import annotations

import logging
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from models import SchedulePageResult

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_THREE_COL_HEADERS = ["Identity", "Description", "Current Value"]
_FOUR_COL_HEADERS = ["Identity", "Description", "Cost", "Current Value"]
_THREE_COL_WIDTHS = [32, 55, 20]
_FOUR_COL_WIDTHS = [30, 45, 16, 20]


def _write_page_sheet(worksheet: Worksheet, page: SchedulePageResult) -> None:
    # Use 'Investment' as the identity header for tickered layouts
    identity_label = "Investment" if (page.source or "").upper().startswith("TICKERED") else "Identity"
    if page.has_cost_column:
        headers = [_FOUR_COL_HEADERS[0].replace("Identity", identity_label)] + _FOUR_COL_HEADERS[1:]
    else:
        headers = [_THREE_COL_HEADERS[0].replace("Identity", identity_label)] + _THREE_COL_HEADERS[1:]
    widths = _FOUR_COL_WIDTHS if page.has_cost_column else _THREE_COL_WIDTHS
    numeric_column_indices = {len(headers) - 1, len(headers)} if page.has_cost_column else {len(headers)}

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    for row_idx, row in enumerate(page.rows, start=2):
        for col_idx, value in enumerate(row.to_list(include_cost=page.has_cost_column), start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            cell.alignment = _RIGHT if col_idx in numeric_column_indices else _LEFT

    for col_idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(ord("A") + col_idx - 1)].width = width


def _unique_sheet_name(desired: str, used_names: set) -> str:
    """Excel worksheet names must be unique (and <= 31 chars). Page numbers
    already satisfy the length limit, but guard uniqueness defensively.
    """
    name = desired[:31]
    suffix = 1
    while name in used_names:
        suffix += 1
        candidate = f"{desired}_{suffix}"
        name = candidate[:31]
    return name


def write_workbook(pages: List[SchedulePageResult], output_path: str) -> None:
    """Write one worksheet per non-empty Schedule H page to `output_path`."""
    non_empty_pages = [page for page in pages if not page.is_empty()]
    if not non_empty_pages:
        raise ValueError("No non-empty Schedule H pages to write.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set = set()
    for page in non_empty_pages:
        sheet_name = _unique_sheet_name(str(page.page_number), used_names)
        used_names.add(sheet_name)
        worksheet = workbook.create_sheet(title=sheet_name)
        _write_page_sheet(worksheet, page)

    workbook.save(output_path)
    logger.info("Workbook written: %s", output_path)


def write_error_workbook(output_path: str, error_text: str) -> None:
    """Write a single-sheet workbook containing the error text for failed PDFs."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Error"
    # Write the error text into A1 and enable wrap
    cell = worksheet.cell(row=1, column=1, value=error_text)
    cell.alignment = Alignment(wrap_text=True)
    # widen the column so the text is readable
    worksheet.column_dimensions["A"].width = 120
    workbook.save(output_path)
    logger.info("Error workbook written: %s", output_path)
